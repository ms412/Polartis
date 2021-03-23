import sys
from polatis import Polatis
from configobj import ConfigObj
from openpyxl import load_workbook

class crossconnection(object):

    def __init__(self,configfile):

        self._configfile = configfile

    def readConfig(self):
        _config = ConfigObj(self._configfile)['POLATIS']

        if bool(_config) is False:
            print('ERROR config file not found', self._configfile)
            sys.exit()

        self._switchHost = _config.get('HOST',None)
        self._switchPort = _config.get('PORT',None)
        self._switchUser = _config.get('USER', None)
        self._switchPassword = _config.get('PASSWORD',None)
        return True

    def connect(self):
        self._pol = Polatis()
        self._pol.connect('192.168.2.3',8008,'newton','newton')

    def createCCfromExel(self,file):
        _wb = load_workbook(file)
        _sheet = _wb['CC']
        for row_cells in _sheet.iter_rows(min_row=2):
            print ('CellA= %s ,CellB= %s'% (row_cells[1].value, row_cells[2].value))
            self._pol.createCrossConnection(row_cells[0],row_cells[2],False)

    def start(self):
        self.readConfig()
        self.connect()
        self.createCCfromExel('./crossconnection.xlsx')


if __name__ == '__main__':
    cc = crossconnection('./Polatis.config')
    cc.start()
